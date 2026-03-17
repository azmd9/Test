"""Store and retrieve currency rate data as Excel and JSON."""

import json
import logging
import os
from datetime import datetime, timezone, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

import config

logger = logging.getLogger(__name__)


def _ensure_data_dir():
    """Create the data directory if it doesn't exist."""
    Path(config.DATA_DIR).mkdir(parents=True, exist_ok=True)


def save_rates_json(rate_data: dict) -> str:
    """Append rate data to a JSON Lines file (one JSON object per line).

    Args:
        rate_data: Dict with date, pairs, and sources.

    Returns:
        Path to the written file.
    """
    _ensure_data_dir()
    filepath = os.path.join(config.DATA_DIR, "rates.jsonl")

    with open(filepath, "a") as f:
        f.write(json.dumps(rate_data) + "\n")

    logger.info("Saved JSON rates to %s", filepath)
    return filepath


def _derive_source(pair_key: str) -> str:
    """Derive the data source label from a currency pair key like 'EUR/USD'."""
    frm, to = pair_key.split("/")
    sources = set()
    for ccy in (frm, to):
        if ccy == "EUR":
            continue
        if ccy in config.SECONDARY_ONLY_CURRENCIES:
            sources.add("ExchangeRate-API")
        else:
            sources.add("Frankfurter/ECB")
    return " + ".join(sorted(sources)) if sources else "Frankfurter/ECB"


def save_rates_excel(rate_data: dict) -> str:
    """Append rate data to the cumulative rates.xlsx file.

    Uses a tall format with one row per currency pair per day:
        Rate Date | Currency Pair | Rate (vs EUR) | Source | Retrieved (UTC)

    If the date already exists in the file, its rows are skipped to avoid
    duplicates (allows safe re-runs).

    Args:
        rate_data: Dict with date, pairs, and sources.

    Returns:
        Path to the written file.
    """
    _ensure_data_dir()
    cumulative_path = os.path.join(config.DATA_DIR, "rates.xlsx")
    cet = ZoneInfo("Europe/Berlin")
    retrieved_at = datetime.now(cet).replace(hour=7, minute=30, second=0, microsecond=0).strftime("%Y-%m-%d %H:%M CET")
    headers = ["Rate Date", "Currency Pair", "Rate (vs EUR)", "Source", "Retrieved (CET)"]

    if os.path.exists(cumulative_path):
        wb = load_workbook(cumulative_path)
        ws = wb.active
        # Check if this date already exists (skip duplicates)
        existing_dates = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                existing_dates.add(str(row[0]))
        if rate_data["date"] in existing_dates:
            logger.info("Date %s already in %s, skipping", rate_data["date"], cumulative_path)
            return cumulative_path
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Currency Rates"
        ws.append(headers)

    sources = rate_data.get("sources", {})
    for pair_key in sorted(rate_data["pairs"].keys()):
        # Derive source from currency if not explicitly provided
        source = sources.get(pair_key, "")
        if not source:
            source = _derive_source(pair_key)
        ws.append([
            rate_data["date"],
            pair_key,
            rate_data["pairs"][pair_key],
            source,
            retrieved_at,
        ])

    wb.save(cumulative_path)
    logger.info("Appended %d rows to %s for %s",
                len(rate_data["pairs"]), cumulative_path, rate_data["date"])
    return cumulative_path


def save_daily_snapshot(rate_data: dict) -> str:
    """Save a standalone JSON file for a single day's rates.

    Args:
        rate_data: Dict with date, pairs, and sources.

    Returns:
        Path to the written file.
    """
    _ensure_data_dir()
    filepath = os.path.join(config.DATA_DIR, f"rates_{rate_data['date']}.json")

    with open(filepath, "w") as f:
        json.dump(rate_data, f, indent=2)

    logger.info("Saved daily snapshot to %s", filepath)
    return filepath


def load_all_rates_json() -> list[dict]:
    """Load all rate entries from the JSON Lines file.

    Returns:
        List of rate data dicts.
    """
    filepath = os.path.join(config.DATA_DIR, "rates.jsonl")
    if not os.path.exists(filepath):
        return []

    entries = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if line:
                entries.append(json.loads(line))
    return entries
